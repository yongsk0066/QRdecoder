import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter

def paint_qrcode_to_excel(array, name):
    wb = openpyxl.Workbook()
    ws = wb.get_sheet_by_name("Sheet")

    for i in range(1, len(array)+1):
        ws.row_dimensions[i].height = 6
        for j in range(1, len(array)+1):
            color = '000000' if array[i-1][j-1] else 'ffffff'
            targetFill = PatternFill(start_color=color,
                                    end_color=color,
                                    fill_type='solid')
            ws.cell(row=i, column=j).fill = targetFill
            ws.column_dimensions[get_column_letter(j)].width = 1

    wb.save(f"{name}.xlsx")
    wb.close()

